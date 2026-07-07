"""
Export des rapports de cotisations en PDF ou Excel.
Statistiques globales, par membre (taux, montants), somme totale collectée, etc.
"""
from decimal import Decimal
from collections import defaultdict

from django.db.models import Count

from .models import CotisationMensuelle


def build_hierarchie_data(annee=None, mois=None):
    """
    Synthèse Regroupement → Section → Sous-section → Dahira (montants, % payé,
    nb cotisations, nb membres). Utilisée à la fois par l'action stats_hierarchie
    et par l'export PDF/Excel, pour ne pas dupliquer la logique d'agrégation.

    Une requête par table (pas une par section/sous-section/dahira) : avec les
    vraies données (~25 dahiras), une boucle avec un .filter()/.count() par
    niveau ferait des dizaines de requêtes séquentielles, très lent en pratique
    sur ce réseau.
    """
    from django.contrib.auth import get_user_model
    from apps.organisation.models import Regroupement, Section, SousSection, Dahira

    User = get_user_model()

    qs = CotisationMensuelle.objects.select_related('membre').all()
    if annee:
        qs = qs.filter(annee=int(annee))
    if mois:
        qs = qs.filter(mois=int(mois))

    by_reg, by_sec, by_ss, by_dahira = {}, {}, {}, {}
    for c in qs:
        m = c.membre
        mt = float(c.montant)
        mp = mt if c.statut == 'payee' else 0
        for key, d in [(m.regroupement_id, by_reg), (m.section_id, by_sec), (m.sous_section_id, by_ss), (m.dahira_id, by_dahira)]:
            if key is None:
                continue
            if key not in d:
                d[key] = {'montant_total': 0, 'montant_paye': 0, 'nb': 0}
            d[key]['montant_total'] += mt
            d[key]['montant_paye'] += mp
            d[key]['nb'] += 1

    def pct(mt, mp):
        return round((mp / mt * 100), 2) if mt else 0

    all_dahiras = list(Dahira.objects.all().order_by('nom'))
    all_sous_sections = list(SousSection.objects.select_related('section').order_by('sexe'))
    all_sections = list(Section.objects.all().order_by('nom'))

    nb_membres_par_dahira = {
        row['dahira_id']: row['nb']
        for row in User.objects.filter(is_active=True, dahira_id__isnull=False)
        .values('dahira_id')
        .annotate(nb=Count('id'))
    }

    dahiras_par_sous_section = {}
    for d in all_dahiras:
        dahiras_par_sous_section.setdefault(d.sous_section_id, []).append(d)
    sous_sections_par_section = {}
    for ss in all_sous_sections:
        sous_sections_par_section.setdefault(ss.section_id, []).append(ss)
    sections_par_regroupement = {}
    for s in all_sections:
        sections_par_regroupement.setdefault(s.regroupement_id, []).append(s)

    def build_dahiras(sous_section_id):
        return [
            {
                'id': d.id,
                'nom': d.nom,
                'montant_total': round(by_dahira.get(d.id, {}).get('montant_total', 0), 2),
                'montant_paye': round(by_dahira.get(d.id, {}).get('montant_paye', 0), 2),
                'pct_paye': pct(by_dahira.get(d.id, {}).get('montant_total', 0), by_dahira.get(d.id, {}).get('montant_paye', 0)),
                'nb_cotisations': by_dahira.get(d.id, {}).get('nb', 0),
                'nb_membres': nb_membres_par_dahira.get(d.id, 0),
            }
            for d in dahiras_par_sous_section.get(sous_section_id, [])
        ]

    def build_sous_sections(section_id):
        return [
            {
                'id': ss.id,
                'label': str(ss),
                'montant_total': round(by_ss.get(ss.id, {}).get('montant_total', 0), 2),
                'montant_paye': round(by_ss.get(ss.id, {}).get('montant_paye', 0), 2),
                'pct_paye': pct(by_ss.get(ss.id, {}).get('montant_total', 0), by_ss.get(ss.id, {}).get('montant_paye', 0)),
                'nb_cotisations': by_ss.get(ss.id, {}).get('nb', 0),
                'dahiras': build_dahiras(ss.id),
            }
            for ss in sous_sections_par_section.get(section_id, [])
        ]

    def build_sections(reg_id):
        return [
            {
                'id': s.id,
                'nom': s.nom,
                'montant_total': round(by_sec.get(s.id, {}).get('montant_total', 0), 2),
                'montant_paye': round(by_sec.get(s.id, {}).get('montant_paye', 0), 2),
                'pct_paye': pct(by_sec.get(s.id, {}).get('montant_total', 0), by_sec.get(s.id, {}).get('montant_paye', 0)),
                'nb_cotisations': by_sec.get(s.id, {}).get('nb', 0),
                'sous_sections': build_sous_sections(s.id),
            }
            for s in sections_par_regroupement.get(reg_id, [])
        ]

    return [
        {
            'id': r.id,
            'nom': r.nom,
            'code': getattr(r, 'code', ''),
            'montant_total': round(by_reg.get(r.id, {}).get('montant_total', 0), 2),
            'montant_paye': round(by_reg.get(r.id, {}).get('montant_paye', 0), 2),
            'pct_paye': pct(by_reg.get(r.id, {}).get('montant_total', 0), by_reg.get(r.id, {}).get('montant_paye', 0)),
            'nb_cotisations': by_reg.get(r.id, {}).get('nb', 0),
            'sections': build_sections(r.id),
        }
        for r in Regroupement.objects.all().order_by('nom')
    ]


def _get_queryset(date_debut=None, date_fin=None, annee=None, mois=None):
    """Queryset des cotisations. Filtres optionnels par période."""
    qs = CotisationMensuelle.objects.exclude(statut='annulee').select_related('membre').order_by('annee', 'mois')
    if annee:
        qs = qs.filter(annee=annee)
    if mois:
        qs = qs.filter(mois=mois)
    if date_debut:
        # Filtrer par mois/annee: on considère la date d'échéance ou mois/annee
        from datetime import date
        qs = qs.filter(annee__gte=date_debut.year).exclude(
            annee=date_debut.year, mois__lt=date_debut.month
        )
    if date_fin:
        from datetime import date
        qs = qs.filter(annee__lte=date_fin.year).exclude(
            annee=date_fin.year, mois__gt=date_fin.month
        )
    return qs


def _get_stats_par_membre(qs):
    """Retourne [{membre_id, nom, nb_total, nb_payees, nb_en_attente, nb_retard, montant_total, montant_paye, taux_cotisation}]"""
    from django.db.models import Sum, Q
    membre_data = defaultdict(lambda: {
        'nb_total': 0, 'nb_payees': 0, 'nb_en_attente': 0, 'nb_retard': 0,
        'montant_total': Decimal('0'), 'montant_paye': Decimal('0'), 'nom': ''
    })
    for c in qs:
        mid = c.membre_id
        membre_data[mid]['nom'] = c.membre.get_full_name() if c.membre else f'Membre #{mid}'
        membre_data[mid]['nb_total'] += 1
        membre_data[mid]['montant_total'] += c.montant
        if c.statut == 'payee':
            membre_data[mid]['nb_payees'] += 1
            membre_data[mid]['montant_paye'] += c.montant
        elif c.statut == 'en_attente':
            membre_data[mid]['nb_en_attente'] += 1
        elif c.statut == 'retard':
            membre_data[mid]['nb_retard'] += 1

    result = []
    for mid, d in membre_data.items():
        nb_total = d['nb_total']
        nb_payees = d['nb_payees']
        montant_total = d['montant_total']
        montant_paye = d['montant_paye']
        taux = round(100 * nb_payees / nb_total, 1) if nb_total else 0
        taux_montant = round(float(100 * montant_paye / montant_total), 1) if montant_total else 0
        result.append({
            'membre_id': mid, 'nom': d['nom'],
            'nb_total': nb_total, 'nb_payees': nb_payees,
            'nb_en_attente': d.get('nb_en_attente', 0), 'nb_retard': d.get('nb_retard', 0),
            'montant_total': montant_total, 'montant_paye': montant_paye,
            'montant_restant': montant_total - montant_paye,
            'taux_cotisation': taux, 'taux_montant': taux_montant
        })
    return sorted(result, key=lambda x: (-x['montant_paye'], x['nom']))


def export_rapport_excel(date_debut=None, date_fin=None, annee=None, mois=None):
    """Export Excel des cotisations + feuille statistiques + feuille par membre."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    qs = _get_queryset(date_debut, date_fin, annee, mois)
    stats_membres = _get_stats_par_membre(qs)

    from django.db.models import Sum, Q
    agg = qs.aggregate(
        total_assigne=Sum('montant'),
        total_paye=Sum('montant', filter=Q(statut='payee')),
    )
    montant_total_collecte = agg.get('total_paye') or Decimal('0')
    montant_total_assigne = agg.get('total_assigne') or Decimal('0')
    nb_total = qs.count()
    nb_payees = qs.filter(statut='payee').count()
    nb_en_attente = qs.filter(statut='en_attente').count()
    nb_retard = qs.filter(statut='retard').count()
    taux_paiement = round(100 * nb_payees / nb_total, 1) if nb_total else 0

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Feuille 1 : Statistiques globales
    ws_stats = wb.active
    ws_stats.title = "Statistiques"
    periode_str = "Toutes les cotisations"
    if annee:
        periode_str = f"Année {annee}" + (f" - Mois {mois}" if mois else "")
    elif date_debut and date_fin:
        periode_str = f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"

    ws_stats.cell(row=1, column=1, value="Rapport des cotisations").font = Font(bold=True, size=14)
    ws_stats.cell(row=2, column=1, value=f"Période : {periode_str}")
    row_s = 4
    for label, val in [
        ("Nombre total de cotisations", nb_total),
        ("Cotisations payées", nb_payees),
        ("En attente", nb_en_attente),
        ("En retard", nb_retard),
        ("Taux de paiement (%)", f"{taux_paiement}%"),
        ("Montant total assigné (FCFA)", float(montant_total_assigne)),
        ("Somme totale collectée (FCFA)", float(montant_total_collecte)),
        ("Reste à collecter (FCFA)", float(montant_total_assigne - montant_total_collecte)),
    ]:
        ws_stats.cell(row=row_s, column=1, value=label).font = header_font
        ws_stats.cell(row=row_s, column=2, value=val).border = border
        row_s += 1

    # Feuille 2 : Taux et montants par membre
    ws_membres = wb.create_sheet("Taux par membre", 1)
    h_membres = ['Membre', 'Cotisations totales', 'Payées', 'En attente', 'En retard', 'Montant assigné (FCFA)',
                 'Montant payé (FCFA)', 'Reste (FCFA)', 'Taux cotisation (%)', 'Taux montant (%)']
    for col, h in enumerate(h_membres, 1):
        c = ws_membres.cell(row=1, column=col, value=h)
        c.font = header_font
        c.border = border
    for i, s in enumerate(stats_membres, 2):
        ws_membres.cell(row=i, column=1, value=s['nom']).border = border
        ws_membres.cell(row=i, column=2, value=s['nb_total']).border = border
        ws_membres.cell(row=i, column=3, value=s['nb_payees']).border = border
        ws_membres.cell(row=i, column=4, value=s['nb_en_attente']).border = border
        ws_membres.cell(row=i, column=5, value=s['nb_retard']).border = border
        ws_membres.cell(row=i, column=6, value=float(s['montant_total'])).border = border
        ws_membres.cell(row=i, column=7, value=float(s['montant_paye'])).border = border
        ws_membres.cell(row=i, column=8, value=float(s['montant_restant'])).border = border
        ws_membres.cell(row=i, column=9, value=f"{s['taux_cotisation']}%").border = border
        ws_membres.cell(row=i, column=10, value=f"{s['taux_montant']}%").border = border
    for col in range(1, 11):
        ws_membres.column_dimensions[get_column_letter(col)].width = 18

    # Feuille 3 : Détail des cotisations
    ws = wb.create_sheet("Détail cotisations", 2)
    headers = ['Membre', 'Type', 'Objet', 'Mois', 'Année', 'Montant (FCFA)', 'Statut', 'Date échéance', 'Date paiement', 'Mode paiement']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.border = border
    taux_par_membre = {s['membre_id']: s['taux_cotisation'] for s in stats_membres}
    for i, c in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=c.membre.get_full_name() if c.membre else '').border = border
        ws.cell(row=i, column=2, value=c.get_type_cotisation_display()).border = border
        ws.cell(row=i, column=3, value=c.objet_assignation or '—').border = border
        ws.cell(row=i, column=4, value=c.mois).border = border
        ws.cell(row=i, column=5, value=c.annee).border = border
        ws.cell(row=i, column=6, value=float(c.montant)).border = border
        ws.cell(row=i, column=7, value=c.get_statut_display()).border = border
        ws.cell(row=i, column=8, value=c.date_echeance.strftime('%d/%m/%Y') if c.date_echeance else '').border = border
        ws.cell(row=i, column=9, value=c.date_paiement.strftime('%d/%m/%Y') if c.date_paiement else '').border = border
        mode_disp = {'wave': 'Wave', 'liquide': 'Espèces', 'autre': 'Autre'}.get(c.mode_paiement or 'wave', c.mode_paiement or '')
        ws.cell(row=i, column=10, value=mode_disp).border = border
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_rapport_pdf(date_debut=None, date_fin=None, annee=None, mois=None):
    """Export PDF des cotisations."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return None

    qs = _get_queryset(date_debut, date_fin, annee, mois)
    stats_membres = _get_stats_par_membre(qs)

    from django.db.models import Sum, Q
    agg = qs.aggregate(
        total_assigne=Sum('montant'),
        total_paye=Sum('montant', filter=Q(statut='payee')),
    )
    montant_total_collecte = agg.get('total_paye') or Decimal('0')
    montant_total_assigne = agg.get('total_assigne') or Decimal('0')
    nb_total = qs.count()
    nb_payees = qs.filter(statut='payee').count()
    nb_en_attente = qs.filter(statut='en_attente').count()
    nb_retard = qs.filter(statut='retard').count()
    taux_paiement = round(100 * nb_payees / nb_total, 1) if nb_total else 0

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    periode_str = "Toutes les cotisations"
    if annee:
        periode_str = f"Année {annee}" + (f" - Mois {mois}" if mois else "")
    elif date_debut and date_fin:
        periode_str = f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"

    from utils.pdf_header import build_pdf_header
    elements.extend(build_pdf_header("Rapport des cotisations", periode_str))

    # Statistiques
    stats_data = [
        ['Nombre total cotisations', str(nb_total)],
        ['Payées', str(nb_payees)],
        ['En attente', str(nb_en_attente)],
        ['En retard', str(nb_retard)],
        ['Taux paiement (%)', f'{taux_paiement}%'],
        ['Montant total assigné (FCFA)', f'{float(montant_total_assigne):,.0f}'],
        ['Somme totale collectée (FCFA)', f'{float(montant_total_collecte):,.0f}'],
        ['Reste à collecter (FCFA)', f'{float(montant_total_assigne - montant_total_collecte):,.0f}'],
    ]
    stats_table = Table([['Indicateur', 'Valeur']] + stats_data, colWidths=[6*cm, 4*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F4D71')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph('<b>Statistiques</b>', styles['Heading2']))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.8*cm))

    # Taux par membre
    elements.append(Paragraph('<b>Taux et montants par membre</b>', styles['Heading2']))
    elements.append(Spacer(1, 0.3*cm))
    if stats_membres:
        m_headers = [['Membre', 'Payées', 'Total', 'Montant assigné', 'Montant payé', 'Taux (%)']]
        m_data = [[s['nom'], str(s['nb_payees']), str(s['nb_total']),
                   f"{float(s['montant_total']):,.0f}", f"{float(s['montant_paye']):,.0f}",
                   f"{s['taux_cotisation']}%"] for s in stats_membres]
        mt = Table(m_headers + m_data, colWidths=[5*cm, 2*cm, 2*cm, 3*cm, 3*cm, 2*cm])
        mt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F4D71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(mt)
    else:
        elements.append(Paragraph('Aucune cotisation.', styles['Normal']))

    doc.build(elements)
    buf.seek(0)
    return buf


def export_hierarchie_excel(annee=None, mois=None):
    """Export Excel de la synthèse hiérarchique (Regroupement → Section → Sous-section → Dahira)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    regroupements = build_hierarchie_data(annee, mois)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Synthèse hiérarchique'

    periode_str = f"Année {annee}" + (f" - Mois {mois}" if mois else "") if annee else "Toutes périodes"
    ws.cell(row=1, column=1, value='Synthèse financière hiérarchique — Ahibahil Khadim').font = Font(bold=True, size=14, color='0F4D71')
    ws.cell(row=2, column=1, value=f"Période : {periode_str}")

    headers = ['Niveau', 'Nom', 'Montant assigné (FCFA)', 'Montant payé (FCFA)', '% payé', 'Nb cotisations', 'Nb membres']
    header_row = 4
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F4D71', end_color='0F4D71', fill_type='solid')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = border

    level_fills = {
        'Regroupement': PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid'),
        'Section': PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid'),
        'Sous-section': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'Dahira': PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid'),
    }
    row = header_row + 1
    for reg in regroupements:
        ws.append(['Regroupement', reg['nom'], reg['montant_total'], reg['montant_paye'], f"{reg['pct_paye']}%", reg['nb_cotisations'], ''])
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = level_fills['Regroupement']
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).border = border
        row += 1
        for sec in reg['sections']:
            ws.append(['Section', f"  {sec['nom']}", sec['montant_total'], sec['montant_paye'], f"{sec['pct_paye']}%", sec['nb_cotisations'], ''])
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = level_fills['Section']
                ws.cell(row=row, column=col).border = border
            row += 1
            for ss in sec['sous_sections']:
                ws.append(['Sous-section', f"    {ss['label']}", ss['montant_total'], ss['montant_paye'], f"{ss['pct_paye']}%", ss['nb_cotisations'], ''])
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = level_fills['Sous-section']
                    ws.cell(row=row, column=col).border = border
                row += 1
                for d in ss['dahiras']:
                    ws.append(['Dahira', f"      {d['nom']}", d['montant_total'], d['montant_paye'], f"{d['pct_paye']}%", d['nb_cotisations'], d['nb_membres']])
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = level_fills['Dahira']
                        ws.cell(row=row, column=col).border = border
                    row += 1

    widths = [14, 32, 20, 20, 10, 16, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f'A{header_row + 1}'

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_hierarchie_pdf(annee=None, mois=None):
    """Export PDF de la synthèse hiérarchique (Regroupement → Section → Sous-section → Dahira)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return None

    regroupements = build_hierarchie_data(annee, mois)

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    periode_str = f"Année {annee}" + (f" - Mois {mois}" if mois else "") if annee else "Toutes périodes"
    from utils.pdf_header import build_pdf_header
    elements.extend(build_pdf_header('Synthèse financière hiérarchique', periode_str))

    level_colors = {
        'reg': colors.HexColor('#0F4D71'),
        'sec': colors.HexColor('#D6EAF8'),
        'ss': colors.white,
        'dahira': colors.HexColor('#FCF3CF'),
    }
    rows = [['Nom', 'Montant assigné', 'Montant payé', '% payé', 'Cotisations']]
    row_styles = []
    idx = 1
    for reg in regroupements:
        rows.append([reg['nom'], f"{reg['montant_total']:,.0f}", f"{reg['montant_paye']:,.0f}", f"{reg['pct_paye']}%", str(reg['nb_cotisations'])])
        row_styles.append(('BACKGROUND', (0, idx), (-1, idx), level_colors['reg']))
        row_styles.append(('TEXTCOLOR', (0, idx), (-1, idx), colors.whitesmoke))
        row_styles.append(('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold'))
        idx += 1
        for sec in reg['sections']:
            rows.append([f"    {sec['nom']}", f"{sec['montant_total']:,.0f}", f"{sec['montant_paye']:,.0f}", f"{sec['pct_paye']}%", str(sec['nb_cotisations'])])
            row_styles.append(('BACKGROUND', (0, idx), (-1, idx), level_colors['sec']))
            idx += 1
            for ss in sec['sous_sections']:
                rows.append([f"        {ss['label']}", f"{ss['montant_total']:,.0f}", f"{ss['montant_paye']:,.0f}", f"{ss['pct_paye']}%", str(ss['nb_cotisations'])])
                idx += 1
                for d in ss['dahiras']:
                    rows.append([f"            {d['nom']}", f"{d['montant_total']:,.0f}", f"{d['montant_paye']:,.0f}", f"{d['pct_paye']}%", str(d['nb_cotisations'])])
                    row_styles.append(('BACKGROUND', (0, idx), (-1, idx), level_colors['dahira']))
                    idx += 1

    t = Table(rows, colWidths=[7*cm, 3.5*cm, 3.5*cm, 2*cm, 2.7*cm], repeatRows=1)
    base_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2DA9E1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
    ] + row_styles
    t.setStyle(TableStyle(base_style))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return buf
