import unicodedata
from decimal import Decimal, InvalidOperation

import openpyxl
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from apps.organisation.models import Section, SousSection, Dahira

User = get_user_model()

DEFAULT_FILE = settings.BASE_DIR.parent / "TABLEAU RECENSEMENT DES SECTIONS HOMMES . FEMMES (1).xlsx"
SHEET_NAME = "TABLEAU SUIVI FINANCIER 2025"
HEADER_ROW = 12  # ligne contenant N°, PRENOM, NOM, ANNEE, SEXE, TELEPHONE, SECTION, DAHIRA, MONTANT
DEFAULT_PASSWORD = "Ahibahil2026!"
DEFAULT_MONTANT = Decimal("1000.00")


def normalize(value):
    """Normalise une chaîne pour une comparaison insensible à la casse/aux accents."""
    if not value:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(value))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).strip().lower()


class Command(BaseCommand):
    help = "Importe les vrais membres depuis le fichier Excel de recensement des sections."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=str(DEFAULT_FILE), help="Chemin vers le fichier .xlsx")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Prévisualise l'import (comptes qui seraient créés) sans écrire en base.",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Fichier introuvable : {file_path}")

        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Onglet '{SHEET_NAME}' introuvable. Onglets disponibles : {wb.sheetnames}")
        ws = wb[SHEET_NAME]

        rows = list(ws.iter_rows(min_row=HEADER_ROW + 1, max_row=ws.max_row, values_only=True))

        # Section "Dakar" déjà seedée par la migration 0004_seed_ahibahil_structure.
        try:
            section_dakar = Section.objects.get(code="SECTION_DAKAR")
        except Section.DoesNotExist:
            raise CommandError(
                "Section 'SECTION_DAKAR' introuvable. Vérifier que les migrations organisation sont appliquées."
            )

        existing_usernames = set(User.objects.values_list("username", flat=True))
        created_count = 0
        skipped_count = 0
        errors = []
        dahiras_created = 0

        with transaction.atomic():
            sid = transaction.savepoint()
            for row in rows:
                numero, prenom, nom, annee, sexe, telephone, section_nom, dahira_nom, montant = (
                    row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8],
                )
                if not prenom or not nom:
                    continue
                prenom = str(prenom).strip()
                nom = str(nom).strip()
                sexe = (str(sexe).strip().upper() if sexe else "")
                if sexe not in ("M", "F"):
                    errors.append(f"Ligne N°{numero} ({prenom} {nom}) : sexe invalide '{sexe}', ignorée.")
                    skipped_count += 1
                    continue

                # Sous-section Homme/Femme rattachée à la section Dakar.
                ss_sexe = "H" if sexe == "M" else "F"
                sous_section, _ = SousSection.objects.get_or_create(section=section_dakar, sexe=ss_sexe)

                # Recherche insensible à la casse/accents d'une dahira existante, sinon création.
                dahira = None
                if dahira_nom:
                    dahira_nom = str(dahira_nom).strip()
                    normalized_target = normalize(dahira_nom)
                    for candidate in Dahira.objects.filter(sous_section=sous_section):
                        if normalize(candidate.nom) == normalized_target:
                            dahira = candidate
                            break
                    if dahira is None:
                        dahira = Dahira.objects.create(sous_section=sous_section, nom=dahira_nom)
                        dahiras_created += 1

                # Montant de cotisation assigné (défaut si absent/illisible).
                try:
                    montant_decimal = Decimal(str(montant)) if montant is not None else DEFAULT_MONTANT
                except InvalidOperation:
                    montant_decimal = DEFAULT_MONTANT

                # Année de naissance (peut être une chaîne dans le fichier).
                annee_naissance = None
                if annee:
                    try:
                        annee_naissance = int(str(annee).strip())
                    except ValueError:
                        annee_naissance = None

                telephone_clean = str(telephone).strip() if telephone else ""

                # Génération d'un username unique à partir du prénom/nom.
                base_username = slugify(f"{prenom}.{nom}").replace("-", "")[:25] or "membre"
                username = base_username
                suffix = 1
                while username in existing_usernames:
                    suffix += 1
                    username = f"{base_username}{suffix}"
                existing_usernames.add(username)

                if dry_run:
                    created_count += 1
                    continue

                User.objects.create_user(
                    username=username,
                    password=DEFAULT_PASSWORD,
                    first_name=prenom,
                    last_name=nom,
                    role="membre",
                    sexe=sexe,
                    telephone=telephone_clean,
                    annee_naissance=annee_naissance,
                    categorie="professionnel",
                    montant_cotisation=montant_decimal,
                    regroupement=section_dakar.regroupement,
                    section=section_dakar,
                    sous_section=sous_section,
                    dahira=dahira,
                    est_actif=True,
                )
                created_count += 1

            if dry_run:
                transaction.savepoint_rollback(sid)
            else:
                transaction.savepoint_commit(sid)

        self.stdout.write(self.style.SUCCESS(
            f"{'[DRY-RUN] ' if dry_run else ''}Membres {'à créer' if dry_run else 'créés'} : {created_count}"
        ))
        self.stdout.write(f"Dahiras {'à créer' if dry_run else 'créées'} : {dahiras_created}")
        if skipped_count:
            self.stdout.write(self.style.WARNING(f"Lignes ignorées : {skipped_count}"))
        for err in errors:
            self.stdout.write(self.style.WARNING(f"  - {err}"))
        if not dry_run and created_count:
            self.stdout.write(self.style.SUCCESS(
                f"Mot de passe temporaire commun pour tous les comptes importés : '{DEFAULT_PASSWORD}' "
                "(à communiquer aux membres, changement recommandé à la première connexion)."
            ))
