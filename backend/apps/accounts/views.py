from decimal import Decimal

from rest_framework import generics, status, filters
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from django.contrib.auth import get_user_model
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend

from .serializers import (
    UserSerializer, UserMinimalSerializer, UserCreateSerializer, UserMeSerializer,
    BadgeSerializer, AttributionBadgeSerializer, HistoriqueConnexionSerializer, JournalActionSerializer,
)
from .models import AttributionBadge, HistoriqueConnexion, JournalAction
from .permissions import (
    is_super_admin,
    can_view_members,
    can_write_members,
    permissions_summary,
    scope_filter,
    CanViewMembers,
    IsAdminRoleOrStaff,
)
from .audit import log_connexion, log_action

User = get_user_model()


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        request = self.context.get('request')
        username = attrs.get(self.username_field)
        try:
            data = super().validate(attrs)
        except Exception:
            # On ne journalise l'échec que si le nom d'utilisateur correspond à un
            # compte existant (mot de passe erroné) — un nom d'utilisateur inconnu
            # n'a pas de FK valide et n'est pas exploitable pour un audit ciblé.
            if request and username:
                candidate = User.objects.filter(**{self.username_field: username}).first()
                if candidate:
                    log_connexion(request, candidate, succes=False)
            raise
        log_connexion(request, self.user, succes=True)
        user_data = dict(UserMeSerializer(self.user, context={'request': request}).data)
        user_data['permissions'] = permissions_summary(self.user)
        data['user'] = user_data
        return data


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    """
    Auto-inscription publique (formulaire /register, non authentifié) OU création
    d'un membre par un Secrétaire Administratif de Cellule / Super Admin depuis
    Gestion des membres (requête authentifiée). Dans ce second cas, on verrouille le
    rattachement organisationnel et le rôle au périmètre de l'auteur — un Secrétaire
    Administratif de Cellule ne doit jamais pouvoir créer un membre (ou un compte à
    privilèges) en dehors de sa propre cellule.
    """
    data = request.data
    requester = request.user
    if requester and requester.is_authenticated and not is_super_admin(requester):
        if getattr(requester, 'role', None) != 'cellule_admin':
            return Response({'detail': "Vous n'avez pas les droits pour créer un membre."}, status=status.HTTP_403_FORBIDDEN)
        data = data.copy()
        data['role'] = 'membre'
        data['regroupement'] = requester.regroupement_id
        data['section'] = requester.section_id
        data['sous_section'] = requester.sous_section_id
        data['dahira'] = requester.dahira_id

    serializer = UserCreateSerializer(data=data)
    if serializer.is_valid():
        user = serializer.save()
        if requester and requester.is_authenticated:
            log_action(
                request, requester, 'membre_create',
                description=f"{user.get_full_name() or user.username} (rôle {user.role})",
                cible_type='membre', cible_id=user.id,
            )
        return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def me(request):
    if request.method == 'GET':
        serializer = UserMeSerializer(request.user, context={'request': request})
        data = dict(serializer.data)
        data['permissions'] = permissions_summary(request.user)
        return Response(data)
    
    # Pour PATCH avec fichier (photo), DRF met les fichiers dans request.data directement
    # ou dans request.FILES selon le parser utilisé
    data = {}
    photo_file = None
    
    for key in request.data:
        value = request.data.get(key)
        if key == 'photo':
            # Vérifie si c'est un fichier (InMemoryUploadedFile ou TemporaryUploadedFile)
            if hasattr(value, 'read'):
                photo_file = value
        else:
            data[key] = value
    
    # Aussi vérifier request.FILES (au cas où)
    if not photo_file and request.FILES.get('photo'):
        photo_file = request.FILES.get('photo')
    
    if photo_file:
        data['photo'] = photo_file
    
    serializer = UserMeSerializer(request.user, data=data, partial=True, context={'request': request})
    if serializer.is_valid():
        user = serializer.save()
        if photo_file:
            from django.utils import timezone
            user.photo_updated_at = timezone.now()
            user.save(update_fields=['photo_updated_at'])
        # Recharger l'utilisateur pour avoir l'URL mise à jour
        user.refresh_from_db()
        data = dict(UserMeSerializer(user, context={'request': request}).data)
        data['permissions'] = permissions_summary(user)
        return Response(data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')
    if not old_password or not new_password:
        return Response(
            {'detail': 'old_password et new_password requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    if not request.user.check_password(old_password):
        return Response(
            {'detail': 'Ancien mot de passe incorrect'},
            status=status.HTTP_400_BAD_REQUEST
        )
    if len(new_password) < 8:
        return Response(
            {'detail': 'Le nouveau mot de passe doit faire au moins 8 caractères'},
            status=status.HTTP_400_BAD_REQUEST
        )
    request.user.set_password(new_password)
    request.user.save(update_fields=['password'])
    return Response({'detail': 'Mot de passe modifié avec succès'})


class UserList(generics.ListAPIView):
    queryset = User.objects.filter(is_active=True).order_by('-date_inscription')
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, CanViewMembers]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = {
        'role': ['exact', 'in'],
        'est_actif': ['exact'],
        'sexe': ['exact'],
        'categorie': ['exact'],
        'groupe_sanguin': ['exact'],
        'profession': ['icontains'],
    }
    search_fields = ['first_name', 'last_name', 'username', 'telephone', 'numero_carte', 'numero_cni']

    def get_serializer_class(self):
        # Utilisé pour peupler des sélecteurs de membres (formulaires) : évite de
        # transférer le profil complet de centaines de membres pour une simple liste
        # déroulante (gain important sur mobile / réseau lent).
        if self.request.query_params.get('minimal') == '1':
            return UserMinimalSerializer
        return UserSerializer

    def get_queryset(self):
        if self.request.query_params.get('minimal') == '1':
            qs = User.objects.filter(is_active=True).only(
                'id', 'username', 'first_name', 'last_name', 'sexe', 'role',
                'regroupement_id', 'section_id', 'sous_section_id', 'dahira_id', 'is_active',
            ).order_by('-date_inscription')
        else:
            # sous_section__section : UserSerializer.get_sous_section_label fait str(sous_section),
            # qui accède à self.section.nom (SousSection.__str__) — sans ce select_related,
            # chaque ligne de la liste déclenche sa propre requête (jusqu'à ~900 aujourd'hui).
            qs = User.objects.filter(is_active=True).select_related(
                'regroupement', 'section', 'sous_section', 'sous_section__section', 'dahira'
            ).order_by('-date_inscription')
        # Périmètre pilote (audit) : chaque rôle scopé ne voit que sa cellule, sa
        # section (cellules pilotes) ou l'ensemble des cellules pilotes (national).
        qs = scope_filter(qs, self.request.user, 'dahira', 'section')

        dahira = self.request.query_params.get('dahira')
        sous_section = self.request.query_params.get('sous_section')
        section = self.request.query_params.get('section')
        regroupement = self.request.query_params.get('regroupement')
        if regroupement:
            qs = qs.filter(regroupement_id=regroupement)
        if section:
            qs = qs.filter(section_id=section)
        if sous_section:
            qs = qs.filter(sous_section_id=sous_section)
        if dahira:
            qs = qs.filter(dahira_id=dahira)
        return qs


class UserDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        """Un utilisateur peut toujours consulter/modifier sa propre fiche (ex. Mon
        profil) ; l'accès à la fiche d'un AUTRE membre suit la matrice de périmètre
        (can_view_members en lecture, can_write_members en écriture)."""
        from rest_framework.exceptions import PermissionDenied

        instance = super().get_object()
        user = self.request.user
        if instance.pk == user.pk:
            return instance
        if self.request.method in ('PUT', 'PATCH', 'DELETE'):
            if not can_write_members(user, dahira_id=instance.dahira_id):
                raise PermissionDenied("Vous ne pouvez modifier que les membres de votre propre cellule.")
        elif not (is_super_admin(user) or can_view_members(user, dahira_id=instance.dahira_id, section_id=instance.section_id)):
            raise PermissionDenied("Ce membre n'est pas dans votre périmètre.")
        return instance

    def _safe_data(self, request, instance):
        """Un Secrétaire Administratif de Cellule (non Super Admin) modifiant un AUTRE
        membre ne doit jamais pouvoir changer son rôle ni le déplacer hors de sa propre
        cellule via le corps de la requête (même s'il a le droit d'éditer sa fiche) —
        même verrou que dans register(). Retourne une copie mutable du payload."""
        data = request.data.copy()
        user = request.user
        if is_super_admin(user) or instance.pk == user.pk:
            return data
        for field in ('role', 'regroupement', 'section', 'sous_section', 'dahira', 'is_staff', 'is_superuser'):
            data.pop(field, None)
        return data

    def _update_and_log(self, request, partial):
        instance = self.get_object()
        old_role = instance.role
        old_name = instance.get_full_name() or instance.username
        data = self._safe_data(request, instance)
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        instance.refresh_from_db()
        if instance.pk != request.user.pk:
            if 'role' in data and instance.role != old_role:
                log_action(
                    request, request.user, 'role_change',
                    description=f"{old_name} : {old_role} → {instance.role}",
                    cible_type='membre', cible_id=instance.id,
                )
            else:
                log_action(
                    request, request.user, 'membre_update',
                    description=old_name, cible_type='membre', cible_id=instance.id,
                )
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        return self._update_and_log(request, kwargs.pop('partial', False))

    def partial_update(self, request, *args, **kwargs):
        """Override pour s'assurer que la catégorie est bien sauvegardée"""
        return self._update_and_log(request, True)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        name = instance.get_full_name() or instance.username
        instance_id = instance.id
        response = super().destroy(request, *args, **kwargs)
        log_action(request, request.user, 'membre_delete', description=name, cible_type='membre', cible_id=instance_id)
        return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stats_admin(request):
    """
    Statistiques du tableau de bord, scopées au périmètre du rôle (séparation des
    tâches) : un rôle finance ne voit pas les stats effectifs, un rôle membres ne voit
    pas les stats finance, chacun limité à sa cellule/section/aux cellules pilotes.
    """
    from apps.accounts.models import CustomUser
    from apps.finance.models import CotisationMensuelle
    from apps.informations.models import Evenement
    from .permissions import has_members_visibility, has_finance_visibility

    user = request.user
    voit_membres = has_members_visibility(user)
    voit_finance = has_finance_visibility(user)
    if not voit_membres and not voit_finance:
        return Response({'detail': 'Non autorisé'}, status=403)

    now = timezone.now()
    annee = now.year
    mois = now.month
    data = {}

    if voit_membres:
        membres_qs = scope_filter(CustomUser.objects.filter(is_active=True), user, 'dahira', 'section')
        data.update({
            'membres_actifs': membres_qs.filter(est_actif=True).count(),
            'total_membres': membres_qs.count(),
            'membres_hommes': membres_qs.filter(sexe='M').count(),
            'membres_femmes': membres_qs.filter(sexe='F').count(),
            'membres_eleves': membres_qs.filter(categorie='eleve').count(),
            'membres_etudiants': membres_qs.filter(categorie='etudiant').count(),
            'membres_professionnels': membres_qs.filter(categorie='professionnel').count(),
        })

    if voit_finance:
        cot_qs = scope_filter(CotisationMensuelle.objects.all(), user, 'membre__dahira', 'membre__section')
        cotisations_total_ce_mois = cot_qs.filter(annee=annee, mois=mois).count()
        cotisations_payees_ce_mois = cot_qs.filter(annee=annee, mois=mois, statut='payee').count()
        taux_ce_mois = (cotisations_payees_ce_mois / cotisations_total_ce_mois * 100) if cotisations_total_ce_mois else 0.0

        cotisations_total_global = cot_qs.count()
        cotisations_payees_global = cot_qs.filter(statut='payee').count()
        taux_global = (cotisations_payees_global / cotisations_total_global * 100) if cotisations_total_global else 0.0

        assignations_en_cours = cot_qs.filter(type_cotisation='assignation', annee=annee).count()

        data.update({
            'cotisations_payees_ce_mois': cotisations_payees_ce_mois,
            'cotisations_total_ce_mois': cotisations_total_ce_mois,
            'taux_paiement_cotisations_ce_mois': round(taux_ce_mois, 2),
            'cotisations_total_global': cotisations_total_global,
            'cotisations_payees_global': cotisations_payees_global,
            'taux_paiement_cotisations_global': round(taux_global, 2),
            'assignations_annuelles_en_cours': assignations_en_cours,
        })

    data['evenements'] = Evenement.objects.filter(est_publie=True).count()
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mes_badges(request):
    qs = AttributionBadge.objects.filter(user=request.user).select_related('badge')
    serializer = AttributionBadgeSerializer(qs, many=True)
    return Response(serializer.data)


# --- Import Excel des membres (recensement type "TABLEAU RECENSEMENT SECTIONS
# HOMMES & FEMMES") -----------------------------------------------------------

_IMPORT_HEADER_ALIASES = {
    'prenom': {'prenom', 'prénom'},
    'nom': {'nom'},
    'sexe': {'sexe'},
    'telephone': {'telephone', 'téléphone', 'tel', 'tél'},
    'section': {'section'},
    'dahira': {'dahira'},
    'montant': {'montant', 'cotisation', 'montant cotisation'},
    'annee': {'annee', 'année', 'annee naissance', 'année naissance'},
    'profession': {'profession'},
    'carte': {'carte', 'numero carte', 'n° carte'},
    'cni': {'cni', 'numero cni', 'n° cni'},
}


def _normalize_header(value):
    import unicodedata
    if value is None:
        return ''
    s = str(value).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s


def _build_username(prenom, nom, taken):
    from django.utils.text import slugify
    base = slugify(f'{prenom}.{nom}') or 'membre'
    username = base
    i = 1
    while username in taken:
        i += 1
        username = f'{base}{i}'
    taken.add(username)
    return username


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def import_membres_excel(request):
    """
    Import en masse de membres depuis un fichier Excel (.xlsx) — recensement type
    "TABLEAU RECENSEMENT SECTIONS HOMMES & FEMMES" (colonnes PRENOM, NOM, ANNEE,
    SEXE, TELEPHONE, SECTION, DAHIRA, MONTANT ; ordre des colonnes libre, en-têtes
    détectés par nom).

    Périmètre : Super Admin peut importer dans n'importe quelle Section/Dahira déjà
    existante ; un Secrétaire Administratif de Cellule ne peut importer que des lignes
    correspondant à sa propre cellule (les autres sont rejetées) — cohérent avec le
    gel des nouvelles cellules pendant le pilote : on ne crée JAMAIS de nouvelle
    Section/Dahira depuis l'import, seules les lignes qui correspondent à une
    Section/Dahira déjà existante sont importées.
    """
    import openpyxl
    from django.contrib.auth.hashers import make_password
    from django.utils.crypto import get_random_string
    from apps.organisation.models import Section, Dahira

    user = request.user
    is_cellule_admin = getattr(user, 'role', None) == 'cellule_admin'
    if not (is_super_admin(user) or is_cellule_admin):
        return Response({'detail': "Vous n'avez pas les droits pour importer des membres."}, status=status.HTTP_403_FORBIDDEN)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return Response({'detail': 'Fichier Excel (.xlsx) requis (champ "file").'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True, read_only=True)
    except Exception:
        return Response({'detail': "Fichier illisible — un .xlsx valide est attendu."}, status=status.HTTP_400_BAD_REQUEST)

    # Index Section/Dahira existantes (jamais de création depuis l'import — gel des
    # cellules hors périmètre pilote).
    sections_par_nom = {s.nom.strip().lower(): s for s in Section.objects.all()}
    dahiras_par_section_et_nom = {}
    for d in Dahira.objects.select_related('sous_section__section'):
        section_id = d.sous_section.section_id if d.sous_section_id else None
        dahiras_par_section_et_nom[(section_id, d.nom.strip().lower())] = d

    existing_usernames = set(User.objects.values_list('username', flat=True))
    existing_pairs = set(
        (fn.strip().lower(), ln.strip().lower(), dahira_id)
        for fn, ln, dahira_id in User.objects.values_list('first_name', 'last_name', 'dahira_id')
    )

    created, already_existing, errors = [], 0, []

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), values_only=True)
        header_row_idx, col = None, {}
        for idx, row in enumerate(rows_iter, start=1):
            normalized = [_normalize_header(c) for c in row]
            found = {}
            for key, aliases in _IMPORT_HEADER_ALIASES.items():
                for pos, cell in enumerate(normalized):
                    if cell in aliases:
                        found[key] = pos
                        break
            if 'prenom' in found and 'nom' in found:
                header_row_idx, col = idx, found
                break
        if header_row_idx is None:
            continue  # feuille sans tableau de membres reconnaissable — ignorée

        for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            def cell(key):
                pos = col.get(key)
                return row[pos] if pos is not None and pos < len(row) else None

            prenom = str(cell('prenom') or '').strip()
            nom = str(cell('nom') or '').strip()
            if not prenom or not nom:
                continue  # ligne vide (fin de tableau ou séparateur)

            section_nom = str(cell('section') or '').strip()
            dahira_nom = str(cell('dahira') or '').strip()
            section = sections_par_nom.get(section_nom.lower())
            dahira = dahiras_par_section_et_nom.get((section.id if section else None, dahira_nom.lower()))

            if not section or not dahira:
                errors.append({'ligne': row_num, 'raison': f"Section/Dahira introuvable : « {section_nom} / {dahira_nom} » (créez-la d'abord dans Organisation)."})
                continue
            if is_cellule_admin and user.dahira_id != dahira.id:
                errors.append({'ligne': row_num, 'raison': f"« {dahira_nom} » n'est pas votre cellule — ligne ignorée."})
                continue

            if (prenom.lower(), nom.lower(), dahira.id) in existing_pairs:
                already_existing += 1
                continue

            sexe_raw = str(cell('sexe') or '').strip().upper()
            sexe = 'M' if sexe_raw.startswith('M') else 'F' if sexe_raw.startswith('F') else ''
            telephone = str(cell('telephone') or '').strip()
            profession = str(cell('profession') or '').strip()
            numero_carte = str(cell('carte') or '').strip()
            numero_cni = str(cell('cni') or '').strip()

            annee_naissance = None
            try:
                annee_val = cell('annee')
                if annee_val not in (None, ''):
                    annee_naissance = int(float(str(annee_val).strip()))
            except (TypeError, ValueError):
                annee_naissance = None

            try:
                montant_val = cell('montant')
                montant = Decimal(str(montant_val).strip()) if montant_val not in (None, '') else Decimal('1000.00')
            except Exception:
                montant = Decimal('1000.00')

            username = _build_username(prenom, nom, existing_usernames)
            new_user = User(
                username=username,
                first_name=prenom,
                last_name=nom,
                sexe=sexe,
                telephone=telephone,
                profession=profession,
                numero_carte=numero_carte,
                numero_cni=numero_cni,
                annee_naissance=annee_naissance,
                montant_cotisation=montant,
                role='membre',
                regroupement_id=section.regroupement_id,
                section=section,
                sous_section=dahira.sous_section,
                dahira=dahira,
                password=make_password(get_random_string(16)),
            )
            created.append(new_user)
            existing_pairs.add((prenom.lower(), nom.lower(), dahira.id))

    User.objects.bulk_create(created, batch_size=200)

    log_action(
        request, user, 'import_excel',
        description=f"{len(created)} créé(s), {already_existing} déjà existant(s), {len(errors)} ligne(s) ignorée(s) — fichier « {uploaded.name} »",
    )

    return Response({
        'created': len(created),
        'already_existing': already_existing,
        'errors': errors[:50],
        'errors_total': len(errors),
    })


class HistoriqueConnexionList(generics.ListAPIView):
    """
    Journal des connexions (page "Logs système", Super Admin uniquement) : qui s'est
    connecté, quand, depuis quelle IP/navigateur/OS, succès ou échec.
    Filtres : ?user=<id>&succes=true|false&adresse_ip=&date_connexion__gte=&date_connexion__lte=
    """
    queryset = HistoriqueConnexion.objects.select_related('user').all()
    serializer_class = HistoriqueConnexionSerializer
    permission_classes = [IsAuthenticated, IsAdminRoleOrStaff]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = {
        'user': ['exact'],
        'succes': ['exact'],
        'adresse_ip': ['exact', 'icontains'],
        'navigateur': ['exact'],
        'systeme_exploitation': ['exact'],
        'date_connexion': ['gte', 'lte'],
    }
    search_fields = ['user__first_name', 'user__last_name', 'user__username', 'adresse_ip']


class JournalActionList(generics.ListAPIView):
    """
    Journal des actions sensibles (page "Logs système", Super Admin uniquement) :
    changements de rôle, création/modification/suppression de membre, validation de
    paiements, import Excel, gel/dégel d'une cellule pilote…
    Filtres : ?acteur=<id>&action=role_change&date_action__gte=&date_action__lte=
    """
    queryset = JournalAction.objects.select_related('acteur').all()
    serializer_class = JournalActionSerializer
    permission_classes = [IsAuthenticated, IsAdminRoleOrStaff]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = {
        'acteur': ['exact'],
        'action': ['exact'],
        'cible_type': ['exact'],
        'adresse_ip': ['exact', 'icontains'],
        'date_action': ['gte', 'lte'],
    }
    search_fields = ['acteur__first_name', 'acteur__last_name', 'acteur__username', 'description']
