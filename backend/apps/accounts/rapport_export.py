"""
Export de la liste des membres en PDF ou Excel (page Gestion des membres).
Respecte les mêmes filtres que /api/auth/users/ (recherche, sexe, catégorie,
profession, groupe sanguin, regroupement/section/dahira) pour que l'export
corresponde à ce que l'admin voit à l'écran.
"""
from django.contrib.auth import get_user_model
from django.db.models import Q

User = get_user_model()

CATEGORIE_LABELS = {'eleve': 'Élève', 'etudiant': 'Étudiant', 'professionnel': 'Professionnel'}
SEXE_LABELS = {'M': 'Masculin', 'F': 'Féminin'}


def get_membres_queryset(params):
    """Reproduit les filtres de UserList.get_queryset à partir d'un dict de query params."""
    qs = User.objects.filter(is_active=True).select_related(
        'regroupement', 'section', 'sous_section', 'dahira'
    ).order_by('section__nom', 'dahira__nom', 'last_name', 'first_name')

    search = params.get('search')
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search) | Q(last_name__icontains=search) |
            Q(username__icontains=search) | Q(telephone__icontains=search) |
            Q(numero_carte__icontains=search) | Q(numero_cni__icontains=search)
        )
    for field in ('sexe', 'categorie', 'groupe_sanguin', 'role', 'est_actif'):
        val = params.get(field)
        if val:
            qs = qs.filter(**{field: val})
    profession = params.get('profession__icontains') or params.get('profession')
    if profession:
        qs = qs.filter(profession__icontains=profession)
    for field, param in (('regroupement_id', 'regroupement'), ('section_id', 'section'), ('dahira_id', 'dahira')):
        val = params.get(param)
        if val:
            qs = qs.filter(**{field: val})
    return qs


def _stats(qs):
    total = qs.count()
    actifs = qs.filter(est_actif=True).count()
    hommes = qs.filter(sexe='M').count()
    femmes = qs.filter(sexe='F').count()
    return {'total': total, 'actifs': actifs, 'hommes': hommes, 'femmes': femmes}


def export_membres_excel(params):
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    qs = get_membres_queryset(params)
    stats = _stats(qs)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F4D71', end_color='0F4D71', fill_type='solid')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_stats = wb.active
    ws_stats.title = 'Statistiques'
    ws_stats.cell(row=1, column=1, value='Rapport des membres — Ahibahil Khadim').font = Font(bold=True, size=14, color='0F4D71')
    ws_stats.cell(row=2, column=1, value=f"{stats['total']} membre(s) correspondant aux filtres appliqués")
    row = 4
    for label, val in [
        ('Total membres (filtré)', stats['total']),
        ('Membres actifs', stats['actifs']),
        ('Hommes', stats['hommes']),
        ('Femmes', stats['femmes']),
    ]:
        ws_stats.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_stats.cell(row=row, column=2, value=val).border = border
        row += 1
    ws_stats.column_dimensions['A'].width = 28
    ws_stats.column_dimensions['B'].width = 14

    ws = wb.create_sheet('Membres', 1)
    headers = ['Nom complet', "Nom d'utilisateur", 'Sexe', 'Téléphone', 'Section', 'Dahira',
               'Catégorie', 'Profession', 'Numéro carte', 'CNI', 'Cotisation mensuelle (FCFA)', 'Statut']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = border
    for i, m in enumerate(qs.iterator(chunk_size=200), 2):
        nom_complet = f'{m.first_name} {m.last_name}'.strip() or m.username
        row_values = [
            nom_complet, m.username, SEXE_LABELS.get(m.sexe, '—'), m.telephone or '—',
            m.section.nom if m.section else '—', m.dahira.nom if m.dahira else '—',
            CATEGORIE_LABELS.get(m.categorie, m.categorie or '—'), m.profession or '—',
            m.numero_carte or '—', m.numero_cni or '—',
            float(m.montant_cotisation) if m.montant_cotisation is not None else 0,
            'Actif' if m.est_actif else 'Inactif',
        ]
        for col, val in enumerate(row_values, 1):
            ws.cell(row=i, column=col, value=val).border = border
    widths = [24, 20, 11, 14, 16, 22, 14, 18, 14, 16, 20, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_membres_pdf(params):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return None

    qs = get_membres_queryset(params)
    stats = _stats(qs)

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    from utils.pdf_header import build_pdf_header
    elements.extend(build_pdf_header('Rapport des membres', f"{stats['total']} membre(s) — filtres appliqués"))

    stats_table = Table([
        ['Total', 'Actifs', 'Hommes', 'Femmes'],
        [str(stats['total']), str(stats['actifs']), str(stats['hommes']), str(stats['femmes'])],
    ], colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F4D71')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.6*cm))

    m_headers = ['Nom complet', 'Sexe', 'Téléphone', 'Section', 'Dahira', 'Carte', 'CNI', 'Cotisation', 'Statut']
    m_rows = [m_headers]
    for m in qs.iterator(chunk_size=200):
        nom_complet = f'{m.first_name} {m.last_name}'.strip() or m.username
        m_rows.append([
            nom_complet, SEXE_LABELS.get(m.sexe, '—'), m.telephone or '—',
            m.section.nom if m.section else '—', m.dahira.nom if m.dahira else '—',
            m.numero_carte or '—', m.numero_cni or '—',
            f'{float(m.montant_cotisation):,.0f}' if m.montant_cotisation is not None else '0',
            'Actif' if m.est_actif else 'Inactif',
        ])
    mt = Table(m_rows, colWidths=[4.2*cm, 2*cm, 3*cm, 3*cm, 3.5*cm, 2.3*cm, 2.8*cm, 2.5*cm, 2*cm], repeatRows=1)
    mt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2DA9E1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4FAFF')]),
    ]))
    elements.append(mt)

    doc.build(elements)
    buf.seek(0)
    return buf
